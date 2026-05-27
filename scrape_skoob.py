import asyncio
import json
import os
import re
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "https://www.skoob.com.br/pt/user/SEU_ID_AQUI/bookshelf"
COOKIES_FILE = "skoob_cookies.json"

async def scrape_page(page, url):
    await page.goto(url, wait_until="domcontentloaded", timeout=60000)
    await asyncio.sleep(8)

    prev_height = 0
    for _ in range(30):
        await page.evaluate("window.scrollBy(0, 800)")
        await asyncio.sleep(0.8)
        height = await page.evaluate("document.body.scrollHeight")
        if height == prev_height:
            break
        prev_height = height

    await asyncio.sleep(3)

    books = []
    seen = set()
    cards = await page.query_selector_all("img[alt^='Capa do livro']")

    for img in cards:
        alt = await img.get_attribute("alt")
        src = await img.get_attribute("src") or ""
        if not alt or alt in seen:
            continue

        raw_title = alt.replace("Capa do livro ", "").strip()
        title = re.sub(r'\s*callback\s*$', '', raw_title, flags=re.IGNORECASE).strip()
        seen.add(alt)

        score = ""
        percentage = ""
        publisher = ""
        author = ""

        try:
            card = await img.evaluate_handle("el => el.parentElement.parentElement")

            author_el = await card.query_selector("h3.text-contrast")
            if author_el:
                author = (await author_el.inner_text()).strip()

            pub_el = await card.query_selector("span.max-w-\\[80px\\]")
            if pub_el:
                publisher = (await pub_el.inner_text()).strip()

            score_els = await card.query_selector_all("span.text-sm.font-bold.text-contrastDark")
            if score_els:
                score = (await score_els[-1].inner_text()).strip()

            pct_els = await card.query_selector_all("span.text-2xs")
            if pct_els:
                percentage = (await pct_els[-1].inner_text()).strip()

        except Exception:
            pass

        books.append({
            "title": title,
            "author": author,
            "publisher": publisher,
            "score": score,
            "percentage": percentage,
            "cover_url": src,
        })

    return books


async def scrape_all():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()

        if os.path.exists(COOKIES_FILE):
            raw = json.load(open(COOKIES_FILE))
            cookies = []
            for c in raw:
                cookie = {
                    "name": c["name"],
                    "value": c["value"],
                    "domain": c.get("domain", ".skoob.com.br"),
                    "path": c.get("path", "/"),
                    "secure": c.get("secure", False),
                    "httpOnly": c.get("httpOnly", False),
                }
                expiry = c.get("expirationDate") or c.get("expires")
                if expiry:
                    cookie["expires"] = int(expiry)
                cookies.append(cookie)
            await context.add_cookies(cookies)
            print(f"Loaded {len(cookies)} cookies")
        else:
            print(f"WARNING: {COOKIES_FILE} not found")

        page = await context.new_page()
        all_books = []
        page_num = 1

        while True:
            url = BASE_URL if page_num == 1 else f"{BASE_URL}?page={page_num}"
            print(f"Scraping page {page_num}: {url}")
            books = await scrape_page(page, url)

            if len(books) == 0 and page_num > 1:
                print(f"  → Got 0, retrying after extra wait...")
                await asyncio.sleep(5)
                books = await scrape_page(page, url)

            print(f"  → Found {len(books)} books")
            all_books.extend(books)

            if len(books) < 30:
                print(f"Last page reached (got {len(books)} < 30). Done.")
                break

            page_num += 1

        await browser.close()
        print(f"\nTotal books collected: {len(all_books)}")
        return all_books


def save_to_excel(books, path="skoob_books.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Minha Estante"

    header_fill = PatternFill("solid", start_color="2E4057")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    headers   = ["#", "Título", "Autor",  "Editora", "Nota", "% Lido", "URL da Capa"]
    col_widths = [5,   50,       30,       20,         8,      10,       80]
    center_cols = {1, 5, 6}

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    row_fill_alt = PatternFill("solid", start_color="F0F4F8")
    data_font = Font(name="Arial", size=10)

    for i, book in enumerate(books, 1):
        row = i + 1
        fill = row_fill_alt if i % 2 == 0 else None
        values = [i, book["title"], book["author"], book["publisher"], book["score"], book["percentage"], book["cover_url"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col in center_cols else "left"
            )
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Saved to {path}")


async def main():
    books = await scrape_all()
    if books:
        save_to_excel(books)
    else:
        print("No books found.")

asyncio.run(main())
